from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl
import re



def clean_string(s):
    return re.sub(r'[\x00-\x1F]', '', s)

# 드라이버 설정
customService = Service(ChromeDriverManager().install())
customOption = Options()
browser = webdriver.Chrome(service = customService, options = customOption)

# 쇼핑 인사이트 접속
URL = 'https://datalab.naver.com/shoppingInsight/sCategory.naver'
browser.get(URL)
browser.implicitly_wait(10)

# 1차 카테고리: 식품 클릭

cateOne = "식품"
cateTwo = ""
cateThird = ""
cateFource = ""



#1차 카테고리 선택
browser.find_element(By.XPATH, '//*[@id="content"]/div[2]/div/div[1]/div/div/div[1]/div/div[1]/span').click()
browser.find_element(By.XPATH, f'//div[2]/div/div[1]/div/div/div[1]/div/div[1]/ul/li/a[contains(text(), "{cateOne}")]').click()


# 2차 카테고리가 비어있지않다면 2차카테고리를 클릭한다.
if cateTwo != '':
    browser.find_element(By.XPATH, '//*[@id="content"]/div[2]/div/div[1]/div/div/div[1]/div/div[2]/span').click()
    browser.find_element(By.XPATH, f'//div[2]/div/div[1]/div/div/div[1]/div/div[2]/ul/li/a[contains(text(), "{cateTwo}")]').click()

# 3차 카테고리가 비어있지않다면 2차카테고리를 클릭한다.
if cateThird != '':
    browser.find_element(By.XPATH, '//*[@id="content"]/div[2]/div/div[1]/div/div/div[1]/div/div[3]/span').click()
    browser.find_element(By.XPATH, f'//div[2]/div/div[1]/div/div/div[1]/div/div[2]/ul/li/a[contains(text(), "{cateThird}")]').click()

# 4차 카테고리가 비어있지않다면 2차카테고리를 클릭한다.
if cateFource != '':
    browser.find_element(By.XPATH, '//*[@id="content"]/div[2]/div/div[1]/div/div/div[1]/div/div[4]/span').click()
    browser.find_element(By.XPATH, f'//div[2]/div/div[1]/div/div/div[1]/div/div[2]/ul/li/a[contains(text(), "{cateFource}")]').click()

## 조회하기 버튼 클릭
browser.find_element(By.XPATH, '//*[@id="content"]/div[2]/div/div[1]/div/a/span').click()
#1초 대기
time.sleep(1)

#엑셀 모듈에서 파일 -> 시트 객체 생성
xlsxFile = openpyxl.Workbook()
xlsxSheet = xlsxFile.active


# 마지막 페이지 번호
# last_page_text = browser.find_element(By.CSS_SELECTOR, "span.page_info").text.strip()  # "1/25"
# last_page = int(last_page_text.split("/")[1])
# print(last_page)


page_info = browser.find_element(By.XPATH, '//span[@class="page_info"]').text.strip()
current_page = int(browser.find_element(By.XPATH, '//span[@class="page_info"]/em').text.strip())
total_page = int(page_info.split("/")[1].strip())

for current_page in range(1 , total_page + 1):
    for j in range(1, 21 , 1):
        keyword = browser.find_element(By.XPATH, f'//*[@id="content"]/div[2]/div/div[2]/div[2]/div/div/div[1]/ul/li[{j}]/a').text
        keyword = keyword.split('\n')
        keyword_clean = clean_string(keyword[1]) if len(keyword) > 1 else clean_string(keyword[0])
        row_num = (current_page - 1) * 20 + j  
        xlsxSheet.cell(row=row_num, column=1).value = row_num
        xlsxSheet.cell(row=row_num, column=2).value = keyword_clean
    if current_page < total_page:
        browser.find_element(By.XPATH, '//*[@id="content"]/div[2]/div/div[2]/div[2]/div/div/div[2]/div/a[2]').click()
        time.sleep(2)
    
#엑셀 파일 저장
xlsxFile.save('result.xlsx')