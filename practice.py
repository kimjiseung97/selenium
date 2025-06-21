#셀레늄 웹드라이버
from selenium import webdriver

#웹드라이버 객체 생성시 수반될 서비스나 옵션
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options

#선택자 및 키보드 입력
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding='utf-8')

# 웹 드라이버 매니저 모듈
from webdriver_manager.chrome import ChromeDriverManager

# 기타모듈
import time

#by 선언
from selenium.webdriver.common.by import By

#서비스 변수 생성 
customService = Service(ChromeDriverManager().install()) 
#옵션 변수 생성 
customOption = Options() 

#드라이버 객체 생성 
browser = webdriver.Chrome(service = customService, options = customOption)

#액셀 모듈 임포트
import openpyxl

# 네이버 바이브 순위차트 크롤링 

URL = "https://vibe.naver.com/chart/total"

browser.get(URL)

browser.implicitly_wait(20)

# 팝업창 x 버튼 /html/body/div[3]/div/div/div/div/a[2]
# x버튼 클릭
browser.find_element(By.XPATH,'/html/body/div[3]/div/div/div/div/a[2]').click()


## 1둥 순위 /html/body/div[1]/div/div[3]/main/div[4]/div[2]/div/table/tbody/tr[1]/td[3]/span

## 1등 가수명 /html/body/div[1]/div/div[3]/main/div[4]/div[2]/div/table/tbody/tr[1]/td[4]/div[2]/span[1]/span/a/span

## 1등 곡 제목 /html/body/div[1]/div/div[3]/main/div[4]/div[2]/div/table/tbody/tr[1]/td[4]/div[1]/span/a/span


## 2등 순위 /html/body/div[1]/div/div[3]/main/div[4]/div[2]/div/table/tbody/tr[2]/td[3]/span

## 2등 가수명 /html/body/div[1]/div/div[3]/main/div[4]/div[2]/div/table/tbody/tr[2]/td[4]/div[2]/span[1]/span/a/span

## 2등 곡 제목 /html/body/div[1]/div/div[3]/main/div[4]/div[2]/div/table/tbody/tr[2]/td[4]/div[1]/span/a/span

## 각 순위별로 어떤값이 바뀌는지 규칙성 파악 


#엑셀파일 생성
xlsxFile = openpyxl.Workbook()

#생성한 파일에서 시트 생성
xlsxSheet = xlsxFile.active


#시트 특정 셀에 데이터 입력
for i in range(10):
    xlsxSheet.cell(row = i + 1, column = 1).value = "hi"
#find_element().text 로 찾은 값을 넣으면 됨


#저장
xlsxFile.save('result.xlsx')


for i in range(1, 101, 1):
    rank = browser.find_element(By.XPATH, f'/html/body/div[1]/div/div[3]/main/div[4]/div[2]/div/table/tbody/tr[{i}]/td[3]/span').text
    title = browser.find_element(By.XPATH, f'/html/body/div[1]/div/div[3]/main/div[4]/div[2]/div/table/tbody/tr[{i}]/td[4]/div[1]/span/a/span').text
    artist = browser.find_element(By.XPATH, f'/html/body/div[1]/div/div[3]/main/div[4]/div[2]/div/table/tbody/tr[{i}]/td[4]/div[2]/span[1]/span/a/span').text
    #셀에 넣는 코드
    xlsxSheet.cell(row = i, column = 1).value = rank
    xlsxSheet.cell(row = i, column = 2).value = title
    xlsxSheet.cell(row = i, column = 3).value = artist    
    print(rank, title, artist)
    
    # print(rank.encode('utf-8').decode('utf-8') , title.encode('utf-8').decode('utf-8'),singer.encode('utf-8').decode('utf-8'))


xlsxFile.save('singer.xlsx')
